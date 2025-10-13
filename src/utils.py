import pandas as pd
from typing import List, Dict, Tuple, Union
import io
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


# Centralized field definitions (matches prompts.py - updated to clinical note structure)
REQUIRED_FIELDS = [
    "Chief_Complaint", "History_Present_Illness", "Past_Medical_History",
    "Current_Medications", "Allergies", "Physical_Exam",
    "Review_of_Systems", "Labs_Imaging_Results", 
    "Assessment_Impression", "Plan"
]


def load_excel_notes(
    file, 
    notes_column: str = "Notes",
    max_rows: int = 10000,
    fill_empty: bool = True
) -> Tuple[List[str], pd.DataFrame]:
    """
    Load clinical notes from Excel file with validation
    
    Args:
        file: Uploaded file object
        notes_column: Name of the column containing notes
        max_rows: Maximum number of rows to process
        fill_empty: If True, replace NaN with empty string instead of dropping
        
    Returns:
        Tuple of (list of clinical notes, original dataframe)
    """
    try:
        # Validate file object
        if file is None:
            raise ValueError("File object is None")
        
        # Try to read the file
        try:
            df = pd.read_excel(file)
        except Exception as read_error:
            # Try CSV as fallback
            try:
                file.seek(0)
                df = pd.read_csv(file)
                print("ℹ️ File read as CSV instead of Excel")
            except:
                raise ValueError(f"Unable to read file as Excel or CSV: {str(read_error)}")
        
        # Check if dataframe is empty
        if df.empty:
            raise ValueError("Excel file is empty (no data rows)")
        
        # Check row limit
        if len(df) > max_rows:
            raise ValueError(
                f"File has {len(df)} rows, which exceeds the maximum of {max_rows}. "
                f"Please split your data into smaller files."
            )
        
        # Check if notes column exists
        if notes_column not in df.columns:
            available_columns = ", ".join(df.columns.tolist())
            raise ValueError(
                f"Column '{notes_column}' not found in Excel file.\n"
                f"Available columns: {available_columns}"
            )
        
        # Handle empty/NaN values - FIXED to maintain alignment
        if fill_empty:
            # Replace NaN with empty string to maintain row alignment
            notes = df[notes_column].fillna("").astype(str).tolist()
            
            # Count empty notes
            empty_count = sum(1 for note in notes if not note.strip())
            if empty_count > 0:
                print(f"⚠️ Warning: {empty_count} empty notes found (will be processed as empty)")
        else:
            # Original behavior: drop NaN (may cause alignment issues)
            notes = df[notes_column].dropna().astype(str).tolist()
            dropped_count = len(df) - len(notes)
            if dropped_count > 0:
                print(f"⚠️ Warning: {dropped_count} rows with empty notes were dropped")
        
        # Check if we have any valid notes
        valid_notes = [note for note in notes if note.strip()]
        if len(valid_notes) == 0:
            raise ValueError(f"No valid notes found in column '{notes_column}' (all empty or NaN)")
        
        # Clean notes (remove excessive whitespace)
        notes = [" ".join(note.split()) for note in notes]
        
        return notes, df
        
    except ValueError as ve:
        raise ve
    except Exception as e:
        raise Exception(f"Error loading Excel file: {str(e)}")


def save_to_excel(
    structured_data: List[Dict], 
    original_df: pd.DataFrame = None,
    sheet_name: str = 'Structured_Features',
    include_original: bool = True,
    add_formatting: bool = True
) -> bytes:
    """
    Save structured data to Excel file with formatting
    
    Args:
        structured_data: List of structured feature dictionaries
        original_df: Original dataframe to merge with (optional)
        sheet_name: Name of the Excel sheet
        include_original: Whether to include original columns
        add_formatting: Whether to add Excel formatting
        
    Returns:
        Bytes object of Excel file
    """
    try:
        # Validate structured data
        if not structured_data:
            raise ValueError("Structured data is empty")
        
        # Convert to DataFrame
        result_df = pd.DataFrame(structured_data)
        
        # Merge with original dataframe if provided
        if original_df is not None and include_original:
            if len(original_df) != len(result_df):
                print(f"⚠️ Warning: Length mismatch - Original: {len(original_df)}, Extracted: {len(result_df)}")
                print(f"   Original data will be saved in separate sheet")
                
                # Save both in separate sheets
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    original_df.to_excel(writer, index=False, sheet_name='Original_Data')
                    result_df.to_excel(writer, index=False, sheet_name='Extracted_Features')
                    
                    if add_formatting:
                        _format_worksheet(writer, 'Original_Data', original_df)
                        _format_worksheet(writer, 'Extracted_Features', result_df)
                
                output.seek(0)
                return output.getvalue()
            
            else:
                # Handle duplicate column names
                original_cols = set(original_df.columns)
                result_cols = set(result_df.columns)
                duplicates = original_cols.intersection(result_cols)
                
                if duplicates:
                    print(f"⚠️ Warning: Duplicate columns found: {duplicates}")
                    # Rename duplicates in result
                    rename_map = {col: f"{col}_extracted" for col in duplicates}
                    result_df = result_df.rename(columns=rename_map)
                
                # Merge horizontally
                result_df = pd.concat([
                    original_df.reset_index(drop=True), 
                    result_df.reset_index(drop=True)
                ], axis=1)
        
        # Save to bytes with formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            if add_formatting:
                _format_worksheet(writer, sheet_name, result_df)
        
        output.seek(0)
        return output.getvalue()
        
    except ValueError as ve:
        raise ve
    except Exception as e:
        raise Exception(f"Error saving to Excel: {str(e)}")


def _format_worksheet(writer, sheet_name: str, df: pd.DataFrame):
    """
    Format Excel worksheet with styling
    
    Args:
        writer: ExcelWriter object
        sheet_name: Name of sheet to format
        df: DataFrame for column sizing
    """
    try:
        worksheet = writer.sheets[sheet_name]
        
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths - FIXED for >26 columns
        for idx, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)  # Works for any number of columns
            
            # Calculate max width
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            adjusted_width = min(max_length + 2, 50)
            
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
        
    except Exception as e:
        print(f"⚠️ Warning: Could not format worksheet: {e}")


def validate_structured_data(data: List[Dict], verbose: bool = False) -> bool:
    """
    Validate that extracted data follows the expected structure
    
    Args:
        data: List of structured dictionaries
        verbose: If True, print detailed validation info
        
    Returns:
        Boolean indicating if data is valid
    """
    # Updated to match new clinical note structure
    required_fields = [
        "Chief_Complaint", "History_Present_Illness", "Past_Medical_History",
        "Current_Medications", "Allergies", "Physical_Exam",
        "Review_of_Systems", "Labs_Imaging_Results", 
        "Assessment_Impression", "Plan"
    ]
    
    # Check if data exists
    if not data or len(data) == 0:
        if verbose:
            print("❌ Validation failed: Data is empty")
        return False
    
    # Check each item
    for idx, item in enumerate(data, 1):
        if not isinstance(item, dict):
            if verbose:
                print(f"❌ Validation failed: Record {idx} is not a dictionary")
            return False
        
        # Check if all required fields are present
        missing_fields = [field for field in required_fields if field not in item]
        if missing_fields:
            if verbose:
                print(f"❌ Validation failed: Record {idx} missing fields: {missing_fields}")
            return False
    
    if verbose:
        print(f"✅ Validation passed: All {len(data)} records have required fields")
    
    return True


def clean_extracted_data(data: List[Dict]) -> List[Dict]:
    """
    Clean and normalize extracted data
    
    Args:
        data: List of structured dictionaries
        
    Returns:
        Cleaned list of structured dictionaries
    """
    cleaned_data = []
    
    for item in data:
        cleaned_item = {}
        
        # Use REQUIRED_FIELDS to ensure consistency
        for field in REQUIRED_FIELDS:
            value = item.get(field, "")
            
            # Convert None to empty string
            if value is None:
                cleaned_item[field] = ""
            # Clean string values
            elif isinstance(value, str):
                # Remove excessive whitespace
                cleaned_item[field] = " ".join(value.split()).strip()
            # Convert other types to string
            else:
                cleaned_item[field] = str(value).strip()
        
        cleaned_data.append(cleaned_item)
    
    return cleaned_data


def get_data_summary(data: List[Dict]) -> Dict:
    """
    Generate summary statistics for extracted data
    
    Args:
        data: List of structured dictionaries
        
    Returns:
        Dictionary with summary statistics
    """
    if not data:
        return {
            "total_records": 0,
            "fields_populated": {},
            "completion_rate": 0.0
        }
    
    total_records = len(data)
    fields_populated = {}
    
    # Count non-empty values for each field
    for field in REQUIRED_FIELDS:
        count = sum(1 for item in data if item.get(field) and str(item.get(field)).strip())
        fields_populated[field] = {
            "count": count,
            "percentage": round((count / total_records * 100), 2)
        }
    
    # Calculate overall completion rate
    total_fields = len(REQUIRED_FIELDS)
    total_populated = sum(item["count"] for item in fields_populated.values())
    completion_rate = (total_populated / (total_records * total_fields) * 100) if total_records > 0 else 0
    
    return {
        "total_records": total_records,
        "fields_populated": fields_populated,
        "completion_rate": round(completion_rate, 2)
    }


def export_to_csv(structured_data: List[Dict], include_bom: bool = True) -> bytes:
    """
    Export structured data to CSV format
    
    Args:
        structured_data: List of structured feature dictionaries
        include_bom: Include UTF-8 BOM for Excel compatibility
        
    Returns:
        Bytes object of CSV file
    """
    try:
        if not structured_data:
            raise ValueError("Cannot export empty data to CSV")
        
        # Convert to DataFrame
        df = pd.DataFrame(structured_data)
        
        # Reorder columns using REQUIRED_FIELDS
        existing_fields = [f for f in REQUIRED_FIELDS if f in df.columns]
        remaining_fields = [f for f in df.columns if f not in existing_fields]
        df = df[existing_fields + remaining_fields]
        
        # Convert to CSV
        csv_string = df.to_csv(index=False)
        
        # Encode with optional BOM for Excel compatibility
        if include_bom:
            return '\ufeff'.encode('utf-8') + csv_string.encode('utf-8')
        else:
            return csv_string.encode('utf-8')
        
    except ValueError as ve:
        raise ve
    except Exception as e:
        raise Exception(f"Error exporting to CSV: {str(e)}")


def export_to_json(structured_data: List[Dict], pretty: bool = True) -> str:
    """
    Export structured data to JSON format
    
    Args:
        structured_data: List of structured feature dictionaries
        pretty: If True, format with indentation
        
    Returns:
        JSON string
    """
    try:
        if not structured_data:
            raise ValueError("Cannot export empty data to JSON")
        
        if pretty:
            return json.dumps(structured_data, indent=2, ensure_ascii=False)
        else:
            return json.dumps(structured_data, ensure_ascii=False)
        
    except (TypeError, ValueError) as e:
        raise ValueError(f"Data is not JSON serializable: {str(e)}")
    except Exception as e:
        raise Exception(f"Error exporting to JSON: {str(e)}")


def filter_empty_records(data: List[Dict]) -> Tuple[List[Dict], int]:
    """
    Filter out records where all fields are empty
    
    Args:
        data: List of structured dictionaries
        
    Returns:
        Tuple of (filtered list, count of removed records)
    """
    if not data:
        return [], 0
    
    filtered_data = []
    removed_count = 0
    
    for item in data:
        # Check if at least one field has a non-empty value
        has_value = any(
            value and str(value).strip() 
            for value in item.values()
        )
        
        if has_value:
            filtered_data.append(item)
        else:
            removed_count += 1
    
    return filtered_data, removed_count


def merge_dataframes(original_df: pd.DataFrame, extracted_df: pd.DataFrame) -> pd.DataFrame:
    """
    Merge original dataframe with extracted features
    
    Args:
        original_df: Original dataframe with clinical notes
        extracted_df: Dataframe with extracted features
        
    Returns:
        Merged dataframe
    """
    try:
        # Ensure both dataframes have the same number of rows
        if len(original_df) != len(extracted_df):
            raise ValueError(
                f"Dataframe length mismatch: original has {len(original_df)} rows, "
                f"extracted has {len(extracted_df)} rows"
            )
        
        # Work on copies to avoid modifying originals
        original_copy = original_df.copy().reset_index(drop=True)
        extracted_copy = extracted_df.copy().reset_index(drop=True)
        
        # Handle duplicate column names
        original_cols = set(original_copy.columns)
        extracted_cols = set(extracted_copy.columns)
        duplicates = original_cols.intersection(extracted_cols)
        
        if duplicates:
            print(f"ℹ️ Found {len(duplicates)} duplicate column(s), renaming...")
            rename_map = {col: f"{col}_extracted" for col in duplicates}
            extracted_copy = extracted_copy.rename(columns=rename_map)
        
        # Concatenate horizontally
        merged_df = pd.concat([original_copy, extracted_copy], axis=1)
        
        return merged_df
        
    except ValueError as ve:
        raise ve
    except Exception as e:
        raise Exception(f"Error merging dataframes: {str(e)}")


def validate_excel_file(file, max_size_mb: int = 50) -> Dict:
    """
    Validate uploaded Excel file
    
    Args:
        file: Uploaded file object
        max_size_mb: Maximum file size in megabytes
        
    Returns:
        Dictionary with validation results
    """
    result = {
        "valid": False,
        "file_type": None,
        "rows": 0,
        "columns": 0,
        "column_names": [],
        "has_data": False,
        "file_size_mb": 0,
        "error": None
    }
    
    try:
        # Check file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        file_size_mb = file_size / (1024 * 1024)
        result["file_size_mb"] = round(file_size_mb, 2)
        
        if file_size_mb > max_size_mb:
            result["error"] = f"File size ({file_size_mb:.2f} MB) exceeds maximum ({max_size_mb} MB)"
            file.seek(0)  # Reset pointer
            return result
        
        # Try reading as Excel
        try:
            df = pd.read_excel(file)
            result["file_type"] = "Excel"
        except:
            # Try CSV as fallback
            file.seek(0)
            try:
                df = pd.read_csv(file)
                result["file_type"] = "CSV"
            except Exception as csv_error:
                result["error"] = f"Unable to read as Excel or CSV: {str(csv_error)}"
                file.seek(0)  # FIXED: Reset pointer
                return result
        
        # FIXED: Reset file pointer for future reads
        file.seek(0)
        
        # Populate validation results
        result["rows"] = len(df)
        result["columns"] = len(df.columns)
        result["column_names"] = df.columns.tolist()
        result["has_data"] = len(df) > 0 and len(df.columns) > 0
        result["valid"] = result["has_data"]
        
        return result
        
    except Exception as e:
        result["error"] = str(e)
        # Ensure file pointer is reset even on error
        try:
            file.seek(0)
        except:
            pass
        return result